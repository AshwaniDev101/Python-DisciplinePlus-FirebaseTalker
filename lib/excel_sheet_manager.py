import pandas as pd
from openpyxl import load_workbook

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class ExcelSheetManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.df = pd.read_excel(file_path)

    def print_data(self):
        print(self.df)
        print("\nColumns:", list(self.df.columns))

    def read_range(self, start_col, end_col, start_row, end_row):
        df = pd.read_excel(
            self.file_path,
            usecols=f"{start_col}:{end_col}",
            nrows=end_row
        )
        return df.iloc[start_row - 1:end_row]

    def read_column_by_letter(self, col_letter):
        col_index = column_index_from_string(col_letter) - 1
        return self.df.iloc[:, col_index]

    def read_column_by_name(self, col_name):
        return self.df[col_name]

    def write_cell(self, cell, value):
        wb = load_workbook(self.file_path)
        sheet = wb.active
        sheet[cell] = value
        wb.save(self.file_path)
        print(f"Written '{value}' to {cell}")

# Example usage:
# manager = ExcelSheetManager(r"S:\Data\temp test game json\xlsx test\example.xlsx")
# manager.print_data()
# print(manager.read_range("B", "I", 1, 8))
# print(manager.read_column("B"))
# manager.write_cell("B27", "Hello World")












# import pandas as pd
# from openpyxl import load_workbook
# Load Excel file
#
# df = pd.read_excel(r"S:\Data\temp test game json\xlsx test\example.xlsx")
#
#
# Load the file
# Load Excel file and print specific data
# df = pd.read_excel(
#     r"S:\Data\temp test game json\xlsx test\example.xlsx",
#     usecols="B:I",     # Columns B to I
#     nrows=8            # Read first 8 rows
# )
#
# print(df)
#
#
# df = pd.read_excel(
#     r"S:\Data\temp test game json\xlsx test\DisciplinePlusData.xlsx",
#     usecols="B",  # Only column B
#     header=None,  # Optional: ignore headers if needed
#     nrows=20
# )
#
# Print the whole DataFrame
# print(df)
#
#
# Save DATA
# # Load workbook and sheet
# file_path = r"S:\Data\temp test game json\xlsx test\DisciplinePlusData.xlsx"
# workbook = load_workbook(file_path)
# sheet = workbook.active  # or workbook['SheetName'] if you know it
#
# # Write to B27
# sheet['B27'] = "Hello World"
#
# # Save workbook
# workbook.save(file_path)
#
