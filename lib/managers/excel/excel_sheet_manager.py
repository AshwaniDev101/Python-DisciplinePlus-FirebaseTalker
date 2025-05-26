import pandas as pd
from openpyxl import load_workbook

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from lib import logger
from lib.managers.json.shared_preferences import SharedPreferences
from lib.models.app_time import AppTime
from lib.models.initiative import Initiative
from lib.models.study_break import StudyBreak



class ExcelSheetManager:

    def __init__(self, file_path):
        self.file_path = file_path
        self.df = pd.read_excel(file_path)



    # Public functions
    def get_initiative_list_from_excel(self, week_name: str):

        pos = self._find_cell_position(week_name)
        df = self._read_range_with_column_number(pos[1], pos[1] + 2, 2, SharedPreferences.get('number_of_rows'))

        # Filter out NaN with is not string but a datatype of panda
        df = df.dropna(subset=[df.columns[0]])

        initList = []

        logger.log(f"\n=== Excel file get [{week_name}] ===")
        for i, (title, duration, break_time) in enumerate(df.itertuples(index=False, name=None)):

            init = Initiative(title=title, completion_time=AppTime(0, duration),study_break=StudyBreak(completion_time=AppTime(0, break_time)), index=i)
            initList.append(init)
            logger.log(f"{init.completion_time} min,\tbrk: {init.study_break.completion_time} min, {init.title} ")

        return initList


    # Private functions
    def _print_data(self):
        logger.log(self.df)
        logger.log(f"\nColumns: {list(self.df.columns)}")

    def _read_range(self, start_col, end_col, start_row, end_row):
        df = pd.read_excel(
            self.file_path,
            usecols=f"{start_col}:{end_col}",
            skiprows=1,
            nrows=end_row - 1
        )
        return df.iloc[start_row - 2:end_row - 1]

    def _read_range_with_column_number(self, start_col_idx, end_col_idx, start_row, end_row):
        df = pd.read_excel(
            self.file_path,
            skiprows=1,
            usecols=lambda col: True  # Load all columns; filter by index below
        )
        return df.iloc[start_row - 2:end_row - 1, start_col_idx:end_col_idx + 1]

    def _find_cell_position(self, target, sheet_name=0):
        df = pd.read_excel(self.file_path, header=None, sheet_name=sheet_name)

        for row_idx in range(df.shape[0]):
            for col_idx in range(df.shape[1]):
                if str(df.iat[row_idx, col_idx]).strip() == target:
                    return row_idx, col_idx  # 0-based index

        return None  # Not found

    def _read_column_by_letter(self, col_letter):
        col_index = column_index_from_string(col_letter) - 1
        return self.df.iloc[:, col_index]

    def _read_column_by_name(self, col_name):
        return self.df[col_name]

    def _write_cell(self, cell, value):
        wb = load_workbook(self.file_path)
        sheet = wb.active
        sheet[cell] = value
        wb.save(self.file_path)
        logger.log(f"Written '{value}' to {cell}")
